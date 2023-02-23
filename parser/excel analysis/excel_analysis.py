import openpyxl as xl
import time


month = ('январь', 'февраль', 'март', 'апрель', 'май', 'июнь',
         'июль', 'август', 'сентябрь', 'октябрь', 'ноябрь', 'декабрь')
days = ('понедельник', 'вторник', 'среда', 'четверг', 'пятница', 'суббота')



def cell_analysis(big_cell):
    """анализирует количество значений разных типов"""
    subject = 0
    teacher = 0
    cabinet = 0
    others_symbol = (' ', '(', ')', '.', ',', '«', '»', '-')

    for value in big_cell:
        small_symbol = len([i for i in str(value) if i.islower()]) <= 1

        if (all((not i.isdigit()) and (i.isupper() or i in others_symbol or small_symbol) for i in str(value)) or all(i in str(value).lower() for i in ('физ', 'куль'))) and value != 'Х':
            subject += 1
        if (len([i for i in str(value) if i.isdigit()]) == 0 and 0 < len([i for i in str(value) if i.isupper()]) <= 3 \
                and len([i for i in str(value) if i.islower()]) >= 3) or any(i in str(value).lower() for i in ('доц', 'ст.преп.')) or value == 'Х':
            teacher += 1
        if len([i for i in str(value) if i.isdigit()]) > 0 and len([i for i in str(value) if i.isupper()]) <= 2 \
                or any(i in str(value).lower() for i in ('улк', 'гук')):
            cabinet += 1

    if len(big_cell) == 3 and (subject, teacher) == (1, 0)\
        and cell_analysis([big_cell[1][0:big_cell[1].index(',')], big_cell[1][big_cell[1].index(','):]]) == (0, 2, 0):
        teacher = 1

    big_cell_value = (subject, teacher, cabinet)

    return big_cell_value


def group_append(sheet):
    """находим группы из документа"""

    new_groups = []  # Временные список Г
    month_in_dok = []

    global month, days

    """находим строку с группами"""
    flag = False
    for row in sheet.iter_rows():
        if flag: break
        for cell in row:
            if not(cell.value == None) and str(cell.value).lower() in month:
                groups_row = cell.row
                flag = True
                break

    """добавляем группы"""
    for cell in sheet[groups_row]:
        if not(cell.value == None) and cell.value.lower() not in month:

            group_new = []
            group_new.append(cell.value)
            new_groups.append(group_new)

    """определяем положение названия дня недели"""
    day_in_dok = []
    for row in sheet.iter_rows(groups_row + 1, groups_row + 1, 1, 10):
        for cell in row:
            if cell.value != None and str(cell.value).lower() in days:
               day_in_dok.append(cell.value)


    if sheet[groups_row+1][0].value != None and sheet[groups_row+1][0].value.lower() in days:
        day_place_left = True
    elif len(day_in_dok) == 1:
        day_place_left = False
    else:
        day_place_left = None


    for row in sheet.iter_rows(groups_row, groups_row, 1, 10):
        for cell in row:
            if cell.value != None and str(cell.value).lower() in month:
                month_in_dok.append(cell.value)

    return day_place_left, groups_row, month_in_dok, new_groups


def date_analys(day, day_place_left, month_in_dok):
    """определяет даты на которые строит рассписание дня"""
    days_cell = []
    dates = []

    global month

    if day_place_left:
        for row in sheet.iter_rows(day-17, day, 2, len(month_in_dok)+1):
            for cell in row:
                if cell.value != None and str(cell.value).lower() not in month and cell.value != 0:
                    days_cell.append(cell)
    elif month == 'если наткнусть на ошибку, переделать':
        pass
    else:
        for row in sheet.iter_rows(day-17, day, 1, len(month_in_dok)):
            for cell in row:
                if cell.value != None and str(cell.value).lower() not in month and cell.value != 0:
                    days_cell.append(cell)

    for cell in days_cell:
        date = f'{cell.value}.{month.index(month_in_dok[cell.column-1].lower())+1}.{str(time.localtime().tm_year)[-2:]}'
        dates.append(date)

    return dates


def get_value(sheet, big_row, big_column):
    """извлекает значения из больших клеток"""
    big_cell = []
    for row in sheet.iter_rows(big_row - 2, big_row, big_column - 3, big_column):
        for cell in row:
            value = cell.value
            if value != None \
                    and ((not any(i in str(value).lower() for i in ['п/г', 'п/п', ' час', 'лб', 'лаб', ' лек', 'гр '])) or 'части' in str(value).lower())\
                    and (len([i for i in str(value) if i.isdigit()]) <= 5 and len([i for i in str(value) if i == '-']) <= 2):
                big_cell.append(value)
            elif value != None and ',' in str(value) and not(len([i for i in str(value) if i == '.']) >= 2 \
                                       and len([i for i in str(value) if i.isdigit()]) >= 4) and \
                    cell_analysis([str(value)[0:str(value).index(',')], str(value)[str(value).index(','):]]) == (0, 0, 2):
                big_cell.append(value)

    for value in big_cell:
        if type(value) == type('x'):
            big_cell[big_cell.index(value)] = value.replace('\n', '')
        if value == 'ФИО':
            big_cell[big_cell.index(value)] = 'ФИО преподавателя'

    if len(big_cell) == 2 and cell_analysis(big_cell) == (2, 0, 0):
        left_side_length = []
        right_side_length = []

        for row in sheet.iter_rows(big_row - 2, big_row, big_column - 3, big_column - 2):
            for cell in row:
                if cell.value != None:
                    left_side_length.append(cell.value)
        for row in sheet.iter_rows(big_row - 2, big_row, big_column - 1, big_column):
            for cell in row:
                if cell.value != None:
                    right_side_length.append(cell.value)

        if len(left_side_length) == 0 or len(right_side_length) == 0:
            big_cell = [big_cell[0] + big_cell[1]]

    if len(big_cell) == 4 and cell_analysis(big_cell) == (4, 0, 0):
        big_cell = [big_cell[0] + big_cell[2], big_cell[1] + big_cell[3]]

    if len(big_cell) == 3 and cell_analysis(big_cell) == (3, 0, 0):
        """если в БК 3 значение предмета, потом доделать"""
        pass

    return big_cell


def day_timetable_analys(day, day_place_left, month_in_dok, new_groups):
    """анализирует рассписание на  день и создает временный список с группами"""
    day_timetable = [] #временный список Х
    column_buffer = []
    row_buffer = []

    dates = date_analys(day, day_place_left, month_in_dok)

    for i in new_groups:
        one_day_group = []
        group = i[0]
        one_day_group.append(group)
        one_day_group.append([dates])
        day_timetable.append(one_day_group)

    if day_place_left == None:
        left_side = len(month_in_dok) + 5
    else:
        left_side = len(month_in_dok) + 6

    #проходит по большим рядам и большим столбцам
    for big_row in range(day-18, day+1, 3):
        if big_row - 2 >= day - 17:
            for big_column in range(left_side, 11+len(new_groups)*4, 4):

                big_cell = get_value(sheet, big_row, big_column)
                big_cell_value = cell_analysis(big_cell)

                #анализирует значение из клеток
                match sum(big_cell_value):
                    case 0:
                        if len(row_buffer) > 0:
                            row_buffer.append([big_row, big_column])
                        elif len(column_buffer) > 0 and any([i[-2]+3, i[-1]] == [big_row, big_column] for i in column_buffer if len(i) > 2):
                            column_buffer.append([big_row, big_column])

                    case 1:
                        if big_cell_value == (1, 0, 0) and cell_analysis(get_value(sheet, big_row, big_column + 4)) != (0, 1, 0):
                            column_buffer.append(big_cell + [big_row, big_column])
                        elif big_cell_value == (0, 0, 1) and \
                                not any([i[-2]+3, i[-1]] == [big_row, big_column] for i in column_buffer):
                            """обработка лекций"""
                            row_buffer.append(big_cell + [big_row, big_column])
                            subject = [row_buffer[0][0], row_buffer[0][1], big_cell[0],
                                       str((big_row - day + 17) // 3 + 1)]

                            for i in row_buffer[1:-1]:
                                day_timetable[(i[1] - 6) // 4 - 1][1].append(subject)

                            for i in row_buffer[0][3], row_buffer[-1][2]:
                                day_timetable[(i - 6) // 4 - 1][1].append(subject)

                            row_buffer = []

                        """обрабатывает 1 значение в бк лекции"""
                        if big_cell_value == (1, 0, 0) and cell_analysis(get_value(sheet, big_row, big_column + 4)) == (0, 1, 0):
                            row_buffer.append(big_cell + get_value(sheet, big_row, big_column + 4) + [big_row, big_column])


                        """обрабатывает случай, когда  лаба на 3 пары и на каждую БК 1 значение"""
                        if big_cell_value == (0, 0, 1) and len(row_buffer) == 0  \
                            and any([i[-2]+3, i[-1]] == [big_row, big_column] for i in column_buffer):
                            for group in column_buffer:
                                if [group[1]+6, group[2]] == [big_row, big_column]:
                                    group.append(big_cell)
                        elif big_cell_value == (0, 1, 0) and len(row_buffer) == 0 \
                                and any([i[-2]+6, i[-1]] == [big_row, big_column] for i in column_buffer):
                            for group in column_buffer:
                                if [group[1]+6, group[2]] == [big_row, big_column]:
                                    subject = [group[0]] + big_cell + [group[-1]] +\
                                              [f'{(group[1] - day + 17) // 3 + 1}-{(big_row - day + 17) // 3 + 1}']
                                    day_timetable[(big_column - 6) // 4 - 1][1].append(subject)
                                    column_buffer.pop(column_buffer.index(group))

                        """обрабатывает физру для 1 курса фэвт'а"""
                        if big_cell_value == (0, 1, 0):
                            if len(row_buffer) > 0 and cell_analysis(get_value(sheet, big_row, big_column + 4)) == (0, 1, 0) \
                                    and all(i in str(row_buffer[0][0]).lower() for i in ['физ', 'куль']):
                                row_buffer.append(big_cell + [big_row, big_column])
                            elif len(row_buffer) > 0 and cell_analysis(get_value(sheet, big_row, big_column + 4)) != (0, 1, 0) \
                                    and all(i in str(row_buffer[0][0]).lower() for i in ['физ', 'куль']):
                                """обработка физры для ФЭВТ 1 курс"""
                                row_buffer.append(big_cell + [big_row, big_column])

                                for group in row_buffer[1:]:
                                    subject = [row_buffer[0][0], group[0], '(classroom)', str((big_row - day + 17) // 3 + 1)]
                                    day_timetable[(group[2] - 6) // 4 - 1][1].append(subject)

                                subject = [row_buffer[0][0], row_buffer[0][1], '(classroom)', str((big_row - day + 17) // 3 + 1)]
                                day_timetable[(row_buffer[0][3] - 6) // 4 - 1][1].append(subject)
                                row_buffer = []

                    case 2:
                        """обрабатывает начало лекции"""
                        if big_cell_value == (1, 1, 0):
                            row_buffer.append(big_cell + [big_row, big_column])
                        elif len(row_buffer) == 0 and big_cell_value == (0, 1, 1) \
                                and not len([i for i in column_buffer if i[-1] == big_column]) == 2:
                            """обработа лабы c 2 конечными значениями"""
                            for group in column_buffer:
                                if [group[-2]+3, group[-1]] == [big_row, big_column]:
                                    subject = [group[0]] + big_cell + \
                                              [f'{(group[1] - day + 17) // 3 + 1}-{(big_row - day + 17) // 3 + 1}']
                                    day_timetable[(big_column - 6) // 4 - 1][1].append(subject)
                                    column_buffer.pop(column_buffer.index(group))

                            """работает с лабами в 3-х парах, где 2 значения в последней БК"""
                        if len(row_buffer) == 0 and big_cell_value == (0, 1, 1) \
                                and len([i for i in column_buffer if i[-1] == big_column]) == 2:
                            for group in column_buffer:
                                if [group[-2]+6, group[-1]] == [big_row, big_column]:
                                    subject = group[0]
                                    para = (group[-2] - day + 17) // 3 + 1

                            column_buffer = [i for i in column_buffer if
                                             [i[-2], i[-1]] not in [[big_row - 3, big_column],
                                                                    [big_row - 6, big_column]]]

                            subjects = [subject, big_cell[-1], big_cell[-2], f'{para}-{para+2}']
                            day_timetable[(big_column - 6) // 4 - 1][1].append(subjects)


                        """добавляет у обоих гр лаба"""
                        if big_cell_value == (2, 0, 0):
                            column_buffer.append([big_cell[0]] + [big_row, big_column])
                            column_buffer.append([big_cell[1]] + [big_row, big_column])

                        """обрабатывает физру у других факультетов"""
                        if big_cell_value == (0, 1, 1) and len(row_buffer) > 0 \
                            and all(i in str(row_buffer[0][0]).lower().replace(' ', '') for i in ['физ', 'куль']) \
                            and cell_analysis(get_value(sheet, big_row, big_column + 4)) == (0, 1, 1) \
                                and not any([i[-2]+3, i[-1]-4] == [big_row, big_column] for i in column_buffer):
                            row_buffer.append(big_cell + [big_row, big_column])
                        elif len(row_buffer) > 0 and big_cell_value == (0, 1, 1) \
                            and (cell_analysis(get_value(sheet, big_row, big_column + 4)) != (0, 1, 1) \
                                 or any([i[-2]+3, i[-1]-4] == [big_row, big_column] for i in column_buffer)):

                            row_buffer.append(big_cell + [big_row, big_column])

                            for group in row_buffer[1:]:
                                subject = [row_buffer[0][0], group[0], group[1],
                                           str((big_row - day + 17) // 3 + 1)]
                                day_timetable[(group[-1] - 6) // 4 - 1][1].append(subject)

                            subject = [row_buffer[0][0], row_buffer[0][1], row_buffer[0][2],
                                       str((big_row - day + 17) // 3 + 1)]

                            day_timetable[(row_buffer[0][-1] - 6) // 4 - 1][1].append(subject)
                            row_buffer = []

                    case _:
                        if sum(big_cell_value) % 2 == 0 \
                                and not(len([i for i in column_buffer if [i[-2]+3, i[-1]] == [big_row, big_column]]) == 2):
                            """обрабатывает почти все концы лаб и некоторые практики"""

                            if big_cell_value[0] == 0 and all(i % 2 == 0 and i > 0 for i in big_cell_value[1:]) \
                                    and len(row_buffer) == 0:
                                """обрабатывает лабы с нормальными значениями"""
                                for group in column_buffer:
                                    if [group[-2] + 3, group[-1]] == [big_row, big_column]:
                                        subject = [group[0]] + big_cell + \
                                                  [f'{(group[-2] - day + 17) // 3 + 1}-{(big_row - day + 17) // 3 + 1}']
                                        day_timetable[(big_column - 6) // 4 - 1][1].append(subject)
                                        column_buffer.pop(column_buffer.index(group))

                            elif big_cell_value[0] == 1 and big_cell_value[1] % 2 == 0 and big_cell_value[2] % 2 != 0 \
                                    and len(row_buffer) == 0:
                                """обрабатывает практики с НЕнормальными значениями"""
                                subject = big_cell + [str((big_row - day + 17) // 3 + 1)]
                                day_timetable[(big_column - 6) // 4 - 1][1].append(subject)

                        elif len([i for i in column_buffer if [i[-2]+3, i[-1]] == [big_row, big_column]]) != 2:

                            """обрабатывает почти все практики, и некоторые концы лаб"""
                            if ((big_cell_value[0] == 1 and big_cell_value[1] == big_cell_value[2]) \
                                    or big_cell_value == (1, 1, 1)) and not all(i in str(big_cell[0]).lower().replace(' ', '') for i in ['физ', 'куль']):
                                if not any([i[-2]+3, i[-1]] == [big_row, big_column] for i in column_buffer):
                                    subject = big_cell + [str((big_row - day + 17) // 3 + 1)]
                                    day_timetable[(big_column - 6) // 4 - 1][1].append(subject)
                                else:
                                    """обрабатывает случай СНАЧАЛА у 1 гр. лаба потом у 1 гр. лаба у другой пр."""
                                    for group in column_buffer:
                                        if (group[1]+3, group[2]) == (big_row, big_column) \
                                                and sheet[big_row-5][big_column-4].value == None:
                                            subject = [big_cell[0], big_cell[3], big_cell[1]] \
                                                      + [str((big_row - day + 17) // 3 + 1)]
                                            day_timetable[(big_column - 6) // 4 - 1][1].append(subject)
                                            subject = [group[0], big_cell[4], big_cell[2]] + \
                                                      [f'{(group[1] - day + 17) // 3 + 1}-{(big_row - day + 17) // 3 + 1}']
                                            day_timetable[(big_column - 6) // 4 - 1][1].append(subject)
                                            column_buffer.pop(column_buffer.index(group))

                                        elif (group[1]+3, group[2]) == (big_row, big_column):
                                            for group in column_buffer:
                                                if (group[1] + 3, group[2]) == (big_row, big_column) \
                                                        and sheet[big_row - 5][big_column - 3].value == None:
                                                    subject = [big_cell[0], big_cell[4], big_cell[2]] \
                                                              + [str((big_row - day + 17) // 3 + 1)]
                                                    day_timetable[(big_column - 6) // 4 - 1][1].append(subject)
                                                    subject = [group[0], big_cell[3], big_cell[1]] + \
                                                              [f'{(group[1] - day + 17) // 3 + 1}-{(big_row - day + 17) // 3 + 1}']
                                                    day_timetable[(big_column - 6) // 4 - 1][1].append(subject)
                                                    column_buffer.pop(column_buffer.index(group))

                            elif big_cell_value[0] == 0 and big_cell_value[1] % 2 == 0 and big_cell_value[2] % 2 != 0 \
                                    and any([i[-2]+3, i[-1]] == [big_row, big_column] for i in column_buffer):
                                """обрабатывает лабы с НЕнормальными значениями"""

                                for group in column_buffer:
                                    if [group[-2]+3, group[-1]] == [big_row, big_column]:
                                        subject = [group[0]] + big_cell + \
                                                  [f'{(group[1] - day + 17) // 3 + 1}-{(big_row - day + 17) // 3 + 1}']
                                        day_timetable[(big_column - 6) // 4 - 1][1].append(subject)
                                        column_buffer.pop(column_buffer.index(group))

                            """заносит физру у других факультетов"""
                            if big_cell_value == (1, 1, 1) \
                                and all(i in str(big_cell[0]).lower().replace(' ', '') for i in ['физ', 'куль'])\
                                and not any([i[-2]+3, i[-1]-4] == [big_row, big_column] for i in column_buffer):
                                row_buffer.append(big_cell + [big_row, big_column])

                        """Обрабатывает случай с двумя лабами"""
                        if len([i for i in column_buffer if [i[-2]+3, i[-1]] == [big_row, big_column]]) == 2:
                            subjects = [i for i in column_buffer if [i[-2]+3, i[-1]] == [big_row, big_column]]
                            for_one_subject = []
                            for_two_subject = []

                            for row in sheet.iter_rows(big_row - 2, big_row, big_column - 3, big_column):
                                for cell in row:
                                    if cell.value != None:
                                        if (big_column - cell.column) >= 2:
                                            for_one_subject.append(cell.value)
                                        elif (big_column - cell.column) < 2:
                                            for_two_subject.append(cell.value)

                            subject = [subjects[0][0]] + for_one_subject + \
                                      [f'{(subjects[0][1] - day + 17) // 3 + 1}-{(big_row - day + 17) // 3 + 1}']
                            day_timetable[(big_column - 6) // 4 - 1][1].append(subject)
                            subject = [subjects[1][0]] + for_two_subject + \
                                      [f'{(subjects[0][1] - day + 17) // 3 + 1}-{(big_row - day + 17) // 3 + 1}']
                            day_timetable[(big_column - 6) // 4 - 1][1].append(subject)

                            for i in subjects:
                                column_buffer.pop(column_buffer.index(i))

    return day_timetable


def timetable_rework(two_week_timetable):
    pass


def get_timetable(passes):
    time_start = time.time()

    timetable = []
    for path in passes:

        book = xl.open(path, read_only=True, data_only=True)
        sheet = book.active

        day_place_left, groups_row, month_in_dok, new_groups = group_append(sheet)

        two_week_timetable = []
        for day in range(groups_row, 109+groups_row, 18):
            if day-17 > 0:
                two_week_timetable += day_timetable_analys(day, day_place_left, month_in_dok, new_groups)

        for day in range(109+groups_row, 219+groups_row, 18):
            if day - 17 > 109 + groups_row:
                two_week_timetable += day_timetable_analys(day, day_place_left, month_in_dok, new_groups)

    time_stop = time.time()

